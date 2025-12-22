import streamlit as st
import pandas as pd
import numpy as np
import math
import io
import openpyxl
from openpyxl.styles import Font, Alignment
import datetime

# --- 1. 核心工具函数 ---

def generate_8760_month_array():
    """
    精确生成一个包含 8760 小时对应月份（1-12）的 numpy 数组。
    假设数据是标准非闰年 (365天)。
    """
    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    hours_in_month = [d * 24 for d in days_in_month]
    
    month_array = np.zeros(8760, dtype=int)
    current_hour_index = 0
    for month in range(1, 13):
        num_hours = hours_in_month[month - 1]
        if current_hour_index + num_hours > 8760:
             num_hours = 8760 - current_hour_index 
        
        month_array[current_hour_index:current_hour_index + num_hours] = month
        current_hour_index += num_hours
    
    return month_array

# --- 2. 核心计算逻辑 (修改放电逻辑) ---

def calculate_single_case(
    pv_unit_data: np.ndarray,
    wind_unit_data: np.ndarray,
    load_data: np.ndarray,
    pv_capacity_mw: float,
    wind_capacity_mw: float,
    storage_power_mw: float,
    storage_duration_h: float,
    storage_efficiency_p3: float,
    discharge_depth_p2: float,
    peak_valley_map: dict,
    prices: dict,
    month_8760_array: np.ndarray,
    discharge_allowed: dict
) -> dict:
    """
    计算单个储能配置下的能源自用比例、消纳量和电价指标。
    """
    # 转换为总发电量 (kWh/h)
    pv_generation = pv_unit_data * pv_capacity_mw
    wind_generation = wind_unit_data * wind_capacity_mw
    generation_data = pv_generation + wind_generation

    # 将MW转换为kWh
    storage_capacity_max_kwh = storage_power_mw * storage_duration_h * 1000
    max_charge_power_kwh = storage_power_mw * 1000
    max_discharge_power_kwh = storage_power_mw * 1000
    
    # 计算单向效率
    storage_efficiency_single = math.sqrt(storage_efficiency_p3)

    # 初始化储能容量数组
    storage_capacity_arr = np.zeros(8760)

    # 初始化统计变量
    total_consumption_sum = 0.0 
    total_on_grid_sum = 0.0 
    total_charge_loss = 0.0 
    total_discharge_loss = 0.0 

    # 尖峰平谷 统计字典初始化 
    time_period_stats = {
        '尖': {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0},
        '峰': {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0},
        '平': {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0},
        '谷': {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0}
    }
    
    for i in range(8760):
        current_generation = generation_data[i]
        current_load = load_data[i]

        previous_storage_capacity = storage_capacity_arr[i - 1] if i > 0 else 0.0
        
        hour_of_day = i % 24
        month = month_8760_array[i] 
        current_period_type = peak_valley_map.get(f"{hour_of_day}_{month}", '平') 

        # --- 充电逻辑 ---
        storage_charge_in_i = 0.0
        on_grid_i = 0.0
        if current_generation > current_load:
            available_from_generation = current_generation - current_load
            remaining_storage_capacity = storage_capacity_max_kwh - previous_storage_capacity
            
            charge_effective = available_from_generation * storage_efficiency_single 

            storage_charge_in_i = min(
                charge_effective,
                remaining_storage_capacity,
                max_charge_power_kwh
            )
            storage_charge_in_i = max(0, storage_charge_in_i)
            
            storage_charge_source = storage_charge_in_i / storage_efficiency_single
            charge_loss_i = storage_charge_source - storage_charge_in_i
            total_charge_loss += charge_loss_i

            on_grid_i = available_from_generation - storage_charge_source
            on_grid_i = max(0, on_grid_i)

        # --- 放电逻辑 (添加时段判断) ---
        storage_required_discharge_i = 0.0
        storage_discharge_out_i = 0.0 
        
        if current_load > current_generation and discharge_allowed.get(current_period_type, False):
            load_gap = current_load - current_generation
            min_storage_capacity = storage_capacity_max_kwh * (1 - discharge_depth_p2)
            max_discharge_from_current_storage = max(0, previous_storage_capacity - min_storage_capacity)

            required_discharge_for_load = load_gap / storage_efficiency_single
            single_hour_discharge_limit = max_discharge_power_kwh 

            storage_required_discharge_i = min(
                max_discharge_from_current_storage,
                required_discharge_for_load,
                single_hour_discharge_limit
            )
            storage_required_discharge_i = max(0, storage_required_discharge_i)
        
            storage_discharge_out_i = storage_required_discharge_i * storage_efficiency_single
        
            discharge_loss_i = storage_required_discharge_i - storage_discharge_out_i
            total_discharge_loss += discharge_loss_i

        # --- 更新储能容量 ---
        storage_capacity_i = previous_storage_capacity + storage_charge_in_i - storage_required_discharge_i
        storage_capacity_i = max(0, min(storage_capacity_i, storage_capacity_max_kwh))
        storage_capacity_arr[i] = storage_capacity_i
        
        # --- 计算指标 ---
        total_available = current_generation + storage_discharge_out_i
        consumption_i = min(total_available, current_load) 
        total_consumption_sum += consumption_i

        if current_generation <= current_load and total_available > current_load:
             on_grid_i += total_available - current_load

        total_on_grid_sum += on_grid_i

        # --- 尖峰平谷 统计及电费计算 ---
        if current_period_type in time_period_stats:
            stats = time_period_stats[current_period_type]
            
            stats['consumption'] += consumption_i
            stats['on_grid'] += on_grid_i

            stats['consumption_cost_sum'] += consumption_i * prices[current_period_type]['self']
            stats['on_grid_cost_sum'] += on_grid_i * prices[current_period_type]['on_grid']

    final_storage_remaining = storage_capacity_arr[-1]
    total_curtailment_sum = total_charge_loss + total_discharge_loss + final_storage_remaining 

    total_generation_sum = np.sum(generation_data) 
    total_pv_generation = np.sum(pv_generation)
    total_wind_generation = np.sum(wind_generation)

    total_consumption_cost = sum(stats['consumption_cost_sum'] for stats in time_period_stats.values())
    total_on_grid_cost = sum(stats['on_grid_cost_sum'] for stats in time_period_stats.values())
    
    weighted_self_price = total_consumption_cost / total_consumption_sum if total_consumption_sum > 0 else 0.0
    
    total_revenue = total_consumption_cost + total_on_grid_cost + (total_curtailment_sum * prices['Curtailment']) 
    integrated_price = total_revenue / total_generation_sum if total_generation_sum > 0 else 0.0

    return {
        "total_generation_sum": total_generation_sum,
        "total_pv_generation": total_pv_generation,
        "total_wind_generation": total_wind_generation,
        "total_consumption_sum": total_consumption_sum,
        "total_on_grid_sum": total_on_grid_sum,
        "total_curtailment_sum": total_curtailment_sum,
        "weighted_self_price": weighted_self_price,
        "weighted_on_grid_price": total_on_grid_cost / total_on_grid_sum if total_on_grid_sum > 0 else 0.0,
        "integrated_price": integrated_price,
        "time_period_stats": time_period_stats
    }

def perform_batch_calculation(
    pv_unit_data: np.ndarray,
    wind_unit_data: np.ndarray,
    load_data: np.ndarray,
    params: dict,
    month_8760_array: np.ndarray 
) -> list:
    """执行批量计算。"""
    batch_results = []
    
    power_start = params['power_start']
    power_end = params['power_end']
    power_step = params['power_step']
    duration_start = params['duration_start']
    duration_end = params['duration_end']
    duration_step = params['duration_step']
    
    total_load_sum = np.sum(load_data)
    
    if power_end < power_start or duration_end < duration_start or power_step <= 0 or duration_step <= 0:
        return []

    current_power = power_start
    
    while current_power <= power_end + 1e-9: 
        current_duration = duration_start
        while current_duration <= duration_end + 1e-9:
            
            results = calculate_single_case(
                pv_unit_data,
                wind_unit_data,
                load_data,
                params['pv_capacity_mw'],
                params['wind_capacity_mw'],
                current_power,
                current_duration,
                params['efficiency'],
                params['depth'],
                params['peak_valley_map'],
                params['prices'],
                month_8760_array,
                params['discharge_allowed']
            )
            
            storage_capacity_kwh = current_power * current_duration * 1000
            total_generation_sum = results["total_generation_sum"]

            batch_results.append({
                "pv_capacity_mw": params['pv_capacity_mw'],
                "wind_capacity_mw": params['wind_capacity_mw'],
                "total_pv_generation": results["total_pv_generation"],
                "total_wind_generation": results["total_wind_generation"],
                "power_mw": current_power,
                "duration_h": current_duration,
                "capacity_kwh": storage_capacity_kwh,
                "weighted_self_price": results["weighted_self_price"],
                "weighted_on_grid_price": results["weighted_on_grid_price"],
                "integrated_price": results["integrated_price"],
                "total_consumption_sum": results["total_consumption_sum"],
                "total_on_grid_sum": results["total_on_grid_sum"],
                "total_curtailment_sum": results["total_curtailment_sum"],
                "self_use_ratio": (results["total_consumption_sum"] / total_generation_sum * 100) if total_generation_sum > 0 else 0.0,
                "load_use_ratio": (results["total_consumption_sum"] / total_load_sum * 100) if total_load_sum > 0 else 0.0,
                "jian_ratio": (results["time_period_stats"]["尖"]["consumption"] / total_generation_sum * 100) if total_generation_sum > 0 else 0.0,
                "feng_ratio": (results["time_period_stats"]["峰"]["consumption"] / total_generation_sum * 100) if total_generation_sum > 0 else 0.0,
                "ping_ratio": (results["time_period_stats"]["平"]["consumption"] / total_generation_sum * 100) if total_generation_sum > 0 else 0.0,
                "gu_ratio": (results["time_period_stats"]["谷"]["consumption"] / total_generation_sum * 100) if total_generation_sum > 0 else 0.0
            })
            current_duration += duration_step
        current_power += power_step
        
    return batch_results

# --- 3. Excel 导出逻辑 (修正了列名访问 Key) ---

def write_batch_results_to_excel(results: list[dict], params: dict) -> io.BytesIO:
    """将批量计算结果写入Excel文件，返回字节流。"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "批量计算汇总"

    headers = [
        "光伏容量 (MW)", "风电容量 (MW)", "光伏发电量 (kWh)", "风电发电量 (kWh)",
        "储能功率 (MW)", "储能时长 (h)", "储能容量 (kWh)",
        "综合电价 (元/kWh)",
        "消纳总电量 (kWh)", "上网总电量 (kWh)", "折损总电量 (kWh)",
        "自用比例 (%)", "用电比例 (%)",
        "尖消纳 (%)", "峰消纳 (%)", "平消纳 (%)", "谷消纳 (%)"
    ]
    sheet.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    # *** 修正 Key Error 的关键区域 (此处使用的 Key 是 DataFrame 重命名后的中文 Key) ***
    for row_idx, result in enumerate(results, 2):
        
        # 1. 光伏容量 (MW) - 已重命名
        sheet.cell(row=row_idx, column=1, value=f"{result['光伏容量 (MW)']:.4f}") 
        # 2. 风电容量 (MW) - 已重命名
        sheet.cell(row=row_idx, column=2, value=f"{result['风电容量 (MW)']:.4f}")
        # 3. 光伏发电量 (kWh) - 现已重命名
        sheet.cell(row=row_idx, column=3, value=f"{result['光伏发电量 (kWh)']:.0f}") 
        # 4. 风电发电量 (kWh) - 现已重命名
        sheet.cell(row=row_idx, column=4, value=f"{result['风电发电量 (kWh)']:.0f}") 
        # 5. 储能功率 (MW) - 已重命名
        sheet.cell(row=row_idx, column=5, value=f"{result['储能功率 (MW)']:.4f}")
        # 6. 储能时长 (h) - 已重命名
        sheet.cell(row=row_idx, column=6, value=f"{result['储能时长 (h)']:.4f}")
        # 7. 储能容量 (kWh) - 已重命名
        sheet.cell(row=row_idx, column=7, value=f"{result['储能容量 (kWh)']:.0f}")
        # 8. 综合电价 (元/kWh) - 已重命名
        sheet.cell(row=row_idx, column=8, value=f"{result['综合电价']:.4f}")
        # 9. 消纳总电量 (kWh) - 已重命名
        sheet.cell(row=row_idx, column=9, value=f"{result['总消纳量 (kWh)']:.0f}")
        # 10. 上网总电量 (kWh) - 已重命名
        sheet.cell(row=row_idx, column=10, value=f"{result['总上网量 (kWh)']:.0f}")
        # 11. 折损总电量 (kWh) - 已重命名
        sheet.cell(row=row_idx, column=11, value=f"{result['总折损量 (kWh)']:.0f}")
        
        # 12-13 列：新增比例
        sheet.cell(row=row_idx, column=12, value=f"{result['自用比例 (%)']:.2f}")
        sheet.cell(row=row_idx, column=13, value=f"{result['用电比例 (%)']:.2f}")
        
        # 14-17 列：分时段消纳占比 (%)
        sheet.cell(row=row_idx, column=14, value=f"{result['尖消纳 (%)']:.2f}")
        sheet.cell(row=row_idx, column=15, value=f"{result['峰消纳 (%)']:.2f}")
        sheet.cell(row=row_idx, column=16, value=f"{result['平消纳 (%)']:.2f}")
        sheet.cell(row=row_idx, column=17, value=f"{result['谷消纳 (%)']:.2f}")

    current_row = sheet.max_row + 2
    sheet.cell(row=current_row, column=1, value="--- 输入参数 ---").font = Font(bold=True)
    current_row += 1
    sheet.cell(row=current_row, column=1, value="储能往返效率 (P3)").font = Font(bold=True)
    sheet.cell(row=current_row, column=2, value=f"{params['efficiency']:.4f}")
    current_row += 1
    sheet.cell(row=current_row, column=1, value="放电深度 (P2)").font = Font(bold=True)
    sheet.cell(row=current_row, column=2, value=f"{params['depth']:.4f}")
    current_row += 1
    
    sheet.cell(row=current_row, column=1, value="--- 时段电价 (元/kWh) ---").font = Font(bold=True)
    current_row += 1
    price_headers = ["时段", "自用电价 (元/kWh)", "上网电价 (元/kWh)"]
    for i, header in enumerate(price_headers):
        sheet.cell(row=current_row, column=i+1, value=header).font = Font(bold=True)
    current_row += 1
    for period in ['尖', '峰', '平', '谷']:
        sheet.cell(row=current_row, column=1, value=period)
        sheet.cell(row=current_row, column=2, value=f"{params['prices'][period]['self']:.4f}")
        sheet.cell(row=current_row, column=3, value=f"{params['prices'][period]['on_grid']:.4f}")
        current_row += 1
    sheet.cell(row=current_row, column=1, value="折损电价 (元/kWh)").font = Font(bold=True)
    sheet.cell(row=current_row, column=2, value=f"{params['prices']['Curtailment']:.4f}")

    # 新增放电允许设置到Excel
    current_row += 2
    sheet.cell(row=current_row, column=1, value="--- 放电允许设置 ---").font = Font(bold=True)
    current_row += 1
    for period in ['尖', '峰', '平', '谷']:
        sheet.cell(row=current_row, column=1, value=period)
        sheet.cell(row=current_row, column=2, value="允许" if params['discharge_allowed'][period] else "不允许")
        current_row += 1

    current_row += 2
    sheet.cell(row=current_row, column=1, value="--- 尖峰平谷时段配置 (24x12) ---").font = Font(bold=True)
    current_row += 1
    
    pv_df = pd.DataFrame(index=range(24), columns=range(1, 13))
    for hour in range(24):
        for month in range(1, 13):
            pv_df.loc[hour, month] = params['peak_valley_map'].get(f"{hour}_{month}", '平')

    pv_headers = ["小时"] + [f"{m}月" for m in range(1, 13)]
    for col_idx, header in enumerate(pv_headers, 1):
        sheet.cell(row=current_row, column=col_idx, value=header).font = Font(bold=True)
    current_row += 1

    for hour in range(24):
        sheet.cell(row=current_row + hour, column=1, value=f"{hour:02d}:00")
        for month in range(1, 13):
             sheet.cell(row=current_row + hour, column=month + 1, value=pv_df.loc[hour, month])
             
    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream


# --- 4. Streamlit 应用逻辑 (添加了颜色映射逻辑) ---

TIME_PERIODS = ['尖', '峰', '平', '谷']
DEFAULT_MONTHS = list(range(1, 13))

def init_session_state():
    """初始化 Streamlit Session State 中的状态变量"""
    if 'monthly_config_data' not in st.session_state:
        st.session_state.monthly_config_data = {
            m: {'尖': '19-21', '峰': '12-14, 17-18, 22-23', '平': '7-11, 15-16', '谷': '0-6'} for m in DEFAULT_MONTHS
        }
    if 'batch_results_df' not in st.session_state:
        st.session_state.batch_results_df = pd.DataFrame()
    if 'params_for_excel' not in st.session_state:
        st.session_state.params_for_excel = {}
    if 'discharge_allowed' not in st.session_state:
        st.session_state.discharge_allowed = {period: True for period in TIME_PERIODS}

def parse_time_slot_input(input_str):
    """解析时间段输入，返回 (start, end) 列表，或 None 表示格式错误。"""
    slots = []
    parts = input_str.replace(' ', '').split(',')
    for part in parts:
        if not part: continue
        try:
            if '-' not in part: return None 
            start_str, end_str = part.split('-')
            start = int(start_str)
            end = int(end_str)
            if 0 <= start <= 23 and 0 <= end <= 23:
                slots.append((start, end))
            else:
                return None
        except ValueError:
            return None
    return slots

def map_config_to_24_hours(month_config):
    """
    将时间段列表转换为 24 小时的时段映射。
    优先级: 谷 (最低) < 平 < 峰 < 尖 (最高)。
    """
    hour_map = {h: '平' for h in range(24)} 
    
    PERIOD_PRIORITY_ORDER = ['谷', '平', '峰', '尖']
    
    for category in PERIOD_PRIORITY_ORDER:
        slots = month_config.get(category, [])
        
        for start, end in slots:
            if slots is None: continue # 跳过无效解析（虽然在 save_current_month_config 已校验）
            if start <= end:
                hours_to_mark = range(start, end + 1)
            else:
                # 跨天处理，例如 22-2
                hours_to_mark = list(range(start, 24)) + list(range(0, end + 1))

            for hour in hours_to_mark:
                # 只有在更高优先级的时段未设置时才覆盖
                if hour_map[hour] == '平' or PERIOD_PRIORITY_ORDER.index(category) > PERIOD_PRIORITY_ORDER.index(hour_map[hour]):
                     hour_map[hour] = category
    
    return hour_map

def generate_peak_valley_map():
    """生成最终的 8760 映射字典 {hour_month: period}"""
    final_map = {}
    parsed_monthly_config = {}
    
    for month in DEFAULT_MONTHS:
        parsed_monthly_config[month] = {}
        for period in TIME_PERIODS:
            input_str = st.session_state.monthly_config_data[month][period]
            slots = parse_time_slot_input(input_str)
            if slots is None and input_str.strip() != "":
                 return None, f"{month}月 '{period}' 时段格式错误（要求格式：'x-x, y-y'，小时数0-23）。"
            parsed_monthly_config[month][period] = slots

    for month in DEFAULT_MONTHS:
        hour_map = map_config_to_24_hours(parsed_monthly_config[month])
        for hour, period in hour_map.items():
            final_map[f"{hour}_{month}"] = period
            
    return final_map, None

# --- 新增：时段颜色映射函数 ---

def color_time_periods(val):
    """根据时段类型返回 CSS 颜色字符串"""
    color_map = {
        '尖': 'background-color: #ff6347; color: white', # 番茄红
        '峰': 'background-color: #ffd700; color: black', # 金色
        '平': 'background-color: #90ee90; color: black', # 浅绿
        '谷': 'background-color: #add8e6; color: black', # 浅蓝
    }
    return color_map.get(val, '')


def render_config_page():
    """渲染 Streamlit 的时段与电价配置页面"""

    st.header("时段与电价配置")

    # --- 时段电价设置 ---
    st.subheader("时段电价 (元/kWh)")

    # 默认价格 (与原始代码保持一致)
    default_prices = {
        '尖': {'self': 1.1, 'on_grid': 0.35}, 
        '峰': {'self': 1.1, 'on_grid': 0.35}, 
        '平': {'self': 0.8, 'on_grid': 0.35}, 
        '谷': {'self': 0.5, 'on_grid': 0.35}
    }
    
    prices = {}
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**时段**")
        for period in TIME_PERIODS:
            st.write(f"**{period}**")
            prices[period] = {}
        st.write("**折损**")

    with col2:
        st.markdown("**自用电价**")
        for period in TIME_PERIODS:
            prices[period]['self'] = st.number_input(f"自用 {period}", value=st.session_state.get(f"price_{period}_self", default_prices[period]['self']), format="%.4f", key=f"price_{period}_self")

    with col3:
        st.markdown("**上网电价**")
        for period in TIME_PERIODS:
            prices[period]['on_grid'] = st.number_input(f"上网 {period}", value=st.session_state.get(f"price_{period}_on_grid", default_prices[period]['on_grid']), format="%.4f", key=f"price_{period}_on_grid")

    prices['Curtailment'] = st.number_input("折损", value=st.session_state.get("price_curtailment", 0.0), format="%.4f", key="price_curtailment", help="用于计算综合电价，默认0")

    # --- 新增: 放电允许设置 ---
    st.subheader("放电允许设置")
    discharge_allowed = {}
    cols = st.columns(4)
    for i, period in enumerate(TIME_PERIODS):
        with cols[i]:
            discharge_allowed[period] = st.checkbox(f"{period} 时段允许放电", value=st.session_state.discharge_allowed.get(period, True), key=f"discharge_allowed_{period}")
    st.session_state.discharge_allowed = discharge_allowed

    # --- 月度时段配置 ---
    st.subheader("月度时段配置")
    st.markdown("格式: `x-x, y-y` (小时数0-23)。优先级：**谷 < 平 < 峰 < 尖**。")

    current_month = st.selectbox("选择配置月份", DEFAULT_MONTHS, key="current_month_select")
    current_config = st.session_state.monthly_config_data[current_month]
    
    st.info(f"正在配置 **{current_month}** 月时段。请点击 **'保存'** 按钮应用更改。")

    input_cols = st.columns(4)
    inputs = {}
    for i, period in enumerate(TIME_PERIODS):
        inputs[period] = input_cols[i].text_input(f"{period}时段", value=current_config[period], key=f"input_{current_month}_{period}")

    def save_current_month_config():
        is_valid = True
        error_msg = ""
        for period, input_str in inputs.items():
            if input_str.strip() != "" and parse_time_slot_input(input_str) is None:
                is_valid = False
                error_msg = f"{current_month}月 '{period}' 时段格式错误。"
                break
            st.session_state.monthly_config_data[current_month][period] = input_str
            
        if is_valid:
            st.success(f"{current_month} 月配置保存成功！")
        else:
            st.error(f"配置保存失败: {error_msg}")
            
    st.button("保存当前月配置", on_click=save_current_month_config)
    
    with st.expander("批量复制配置"):
        target_months = st.multiselect("选择目标月份", [m for m in DEFAULT_MONTHS if m != current_month], key="copy_months")
        
        def execute_copy_config():
            if not target_months:
                st.warning("请选择目标月份。")
                return
            save_current_month_config() 
            source_config = st.session_state.monthly_config_data[current_month]
            for month in target_months:
                st.session_state.monthly_config_data[month] = source_config.copy() 
            st.success(f"已将 {current_month} 月的配置复制到 {len(target_months)} 个月份。")
            
        st.button("执行复制", on_click=execute_copy_config)
        
    with st.expander("时段映射预览 (24x12)"):
        final_map, error_msg = generate_peak_valley_map()
        if error_msg:
            st.error(f"时段映射生成失败: {error_msg}")
        elif final_map:
            df_preview = pd.DataFrame(index=range(24), columns=range(1, 13))
            for hour in range(24):
                for month in range(1, 13):
                    df_preview.loc[hour, month] = final_map.get(f"{hour}_{month}", '平')
            df_preview.index.name = "小时"
            
            # --- 关键修改：应用颜色样式 ---
            st.markdown("##### 颜色图例: <span style='background-color:#ff6347; color:white; padding: 2px 5px;'>尖</span> <span style='background-color:#ffd700; color:black; padding: 2px 5px;'>峰</span> <span style='background-color:#90ee90; color:black; padding: 2px 5px;'>平</span> <span style='background-color:#add8e6; color:black; padding: 2px 5px;'>谷</span>", unsafe_allow_html=True)
            
            st.dataframe(
                df_preview.style.applymap(color_time_periods), # 应用颜色映射函数
                use_container_width=True
            )
            
    return prices

def render_main_page(prices):
    """渲染 Streamlit 的批量计算与参数设置页面"""
    
    st.title("新能源源荷匹配分析模型")

    # --- 1. 数据输入 ---
    st.subheader("数据文件 (8760小时)")
    uploaded_file = st.file_uploader("上传 CSV 文件 (需包含 PV_Unit_Output(kWh), Wind_Unit_Output(kWh), Load(kWh) 列)", type="csv")
    
    data_loaded = False
    pv_unit_data = None
    wind_unit_data = None
    load_data = None
    
    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file, encoding='utf-8')
            required_cols = ["PV_Unit_Output(kWh)", "Wind_Unit_Output(kWh)", "Load(kWh)"]
            for col in required_cols:
                if col not in df.columns:
                    st.error(f"文件缺少必需的列: {col}")
                    return
            
            # 由于原始数据文件中的列名是 (kWh)，但 core function 中期望的是单位产出率 (MW)
            # 这里的处理方式（直接使用值/容量MW）暗示了输入数据实际上是每小时的单位MW产出率。
            # 为了保持与原代码逻辑一致，这里暂时使用astype(float)。
            pv_unit_data = df["PV_Unit_Output(kWh)"].values.astype(float)
            wind_unit_data = df["Wind_Unit_Output(kWh)"].values.astype(float)
            load_data = df["Load(kWh)"].values.astype(float)
            
            if len(pv_unit_data) != 8760:
                st.error(f"数据行数不为 8760 小时。当前行数: {len(pv_unit_data)}")
                return
            
            st.success("数据读取成功。")
            data_loaded = True
            
        except Exception as e:
            st.error(f"读取 CSV 文件时发生错误: {e}")
            return

    # --- 2. 容量与参数 ---
    st.subheader("容量与参数")
    
    col1, col2 = st.columns(2)
    pv_capacity_mw = col1.number_input("光伏容量 (MW)", value=st.session_state.get('pv_capacity_mw', 5.0), min_value=0.0, format="%.4f", key='pv_capacity_mw')
    wind_capacity_mw = col2.number_input("风电容量 (MW)", value=st.session_state.get('wind_capacity_mw', 3.0), min_value=0.0, format="%.4f", key='wind_capacity_mw')
    
    col3, col4 = st.columns(2)
    efficiency = col3.number_input("储能往返效率 (P3)", value=st.session_state.get('efficiency', 0.85), min_value=0.0, max_value=1.0, format="%.4f", key='efficiency')
    depth = col4.number_input("放电深度 (P2)", value=st.session_state.get('depth', 0.9), min_value=0.0, max_value=1.0, format="%.4f", key='depth')

    # --- 3. 批量计算参数 ---
    st.subheader("批量计算范围")
    st.markdown("**储能功率 (MW) 设置:**")
    col_p1, col_p2, col_p3 = st.columns(3)
    power_start = col_p1.number_input("起始功率 (MW)", value=st.session_state.get('power_start', 1.0), min_value=0.0, format="%.4f", key="power_start")
    power_end = col_p2.number_input("结束功率 (MW)", value=st.session_state.get('power_end', 4.0), min_value=0.0, format="%.4f", key="power_end")
    power_step = col_p3.number_input("功率步长 (MW)", value=st.session_state.get('power_step', 1.0), min_value=0.1, format="%.4f", key="power_step")

    st.markdown("**储能时长 (h) 设置:**")
    col_d1, col_d2, col_d3 = st.columns(3)
    duration_start = col_d1.number_input("起始时长 (h)", value=st.session_state.get('duration_start', 2.0), min_value=0.0, format="%.4f", key="duration_start")
    duration_end = col_d2.number_input("结束时长 (h)", value=st.session_state.get('duration_end', 4.0), min_value=0.0, format="%.4f", key="duration_end")
    duration_step = col_d3.number_input("时长步长 (h)", value=st.session_state.get('duration_step', 1.0), min_value=0.1, format="%.4f", key="duration_step")

    # --- 4. 批量计算执行 ---
    if st.button("执行计算", disabled=not data_loaded):
        if power_end < power_start or duration_end < duration_start or power_step <= 0 or duration_step <= 0:
            st.error("计算范围或步长设置错误，请检查。")
            return
            
        final_map, error_msg = generate_peak_valley_map()
        if error_msg:
            st.error(f"计算失败: 时段配置错误。{error_msg}")
            return

        params = {
            "pv_capacity_mw": pv_capacity_mw, "wind_capacity_mw": wind_capacity_mw, 
            "efficiency": efficiency, "depth": depth,
            "power_start": power_start, "power_end": power_end, "power_step": power_step,
            "duration_start": duration_start, "duration_end": duration_end, "duration_step": duration_step,
            "peak_valley_map": final_map, "prices": prices,
            "discharge_allowed": st.session_state.discharge_allowed
        }
        
        month_8760_array = generate_8760_month_array() 

        with st.spinner('正在计算...'):
            try:
                batch_results = perform_batch_calculation(
                    pv_unit_data,
                    wind_unit_data,
                    load_data,
                    params,
                    month_8760_array
                )
                
                if batch_results:
                    df_results = pd.DataFrame(batch_results)
                    # 重命名列名，这些重命名的列名将作为字典的 Key 传递给导出函数
                    df_results.rename(columns={
                        "power_mw": "储能功率 (MW)",
                        "duration_h": "储能时长 (h)",
                        "capacity_kwh": "储能容量 (kWh)",
                        "weighted_self_price": "加权自用电价",
                        "weighted_on_grid_price": "加权上网电价",
                        "integrated_price": "综合电价",
                        "total_consumption_sum": "总消纳量 (kWh)",
                        "total_on_grid_sum": "总上网量 (kWh)",
                        "total_curtailment_sum": "总折损量 (kWh)",
                        "self_use_ratio": "自用比例 (%)",
                        "load_use_ratio": "用电比例 (%)",
                        "jian_ratio": "尖消纳 (%)",
                        "feng_ratio": "峰消纳 (%)",
                        "ping_ratio": "平消纳 (%)",
                        "gu_ratio": "谷消纳 (%)",
                        "pv_capacity_mw": "光伏容量 (MW)", 
                        "wind_capacity_mw": "风电容量 (MW)",
                        # 补充这两个关键的重命名
                        "total_pv_generation": "光伏发电量 (kWh)", 
                        "total_wind_generation": "风电发电量 (kWh)" 
                    }, inplace=True)
                    
                    df_results["储能容量 (MWh)"] = (df_results["储能容量 (kWh)"] / 1000).round(1)
                    
                    st.session_state.batch_results_df = df_results
                    st.session_state.params_for_excel = params
                    st.success(f"计算完成！共 {len(batch_results)} 种配置。")
                    
                else:
                    st.warning("没有生成任何计算结果，请检查输入范围和步长。")

            except Exception as e:
                st.error(f"计算过程中发生错误: {e}")

    # --- 5. 结果展示与导出 ---
    
    if not st.session_state.batch_results_df.empty:
        st.subheader("计算结果")
        
        display_df = st.session_state.batch_results_df[[
            "光伏容量 (MW)", "风电容量 (MW)", "储能功率 (MW)", "储能时长 (h)", "储能容量 (MWh)",
            "综合电价", 
            "总消纳量 (kWh)", "总上网量 (kWh)", "总折损量 (kWh)",
            "自用比例 (%)", "用电比例 (%)",
            "尖消纳 (%)", "峰消纳 (%)", "平消纳 (%)", "谷消纳 (%)"
        ]].copy()
        
        st.dataframe(
            display_df.style.format({
                "光伏容量 (MW)": "{:.4f}", "风电容量 (MW)": "{:.4f}",
                "储能功率 (MW)": "{:.4f}", "储能时长 (h)": "{:.4f}",
                "储能容量 (MWh)": "{:.1f}",
                "综合电价": "{:.4f}",
                "总消纳量 (kWh)": "{:.0f}", "总上网量 (kWh)": "{:.0f}", "总折损量 (kWh)": "{:.0f}",
                "自用比例 (%)": "{:.2f}", "用电比例 (%)": "{:.2f}",
                "尖消纳 (%)": "{:.2f}", "峰消纳 (%)": "{:.2f}", "平消纳 (%)": "{:.2f}", "谷消纳 (%)": "{:.2f}",
            }), 
            use_container_width=True
        )

        excel_data = write_batch_results_to_excel(st.session_state.batch_results_df.to_dict('records'), st.session_state.params_for_excel)
        st.download_button(
            label="导出 Excel 完整结果",
            data=excel_data,
            file_name=f"能源经济性批量计算结果_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# --- 5. Streamlit 主函数 ---

def main():
    st.set_page_config(layout="wide", page_title="能源经济性批量计算")
    init_session_state()

    selected_page = st.sidebar.radio("选择页面", ["计算与参数设置", "时段与电价配置"])
    
    if selected_page == "时段与电价配置":
        # 重新渲染配置页面 (因为上面一次渲染在 st.empty() 中，且需要更新 state)
        st.session_state.prices = render_config_page()
    else:
        # 确保在切换到主页时，如果配置页面的价格未更新，使用最新的 session state 价格
        render_main_page(st.session_state.get('prices', {'尖': {'self': 1.1, 'on_grid': 0.35}, '峰': {'self': 1.1, 'on_grid': 0.35}, '平': {'self': 0.8, 'on_grid': 0.35}, '谷': {'self': 0.5, 'on_grid': 0.35}, 'Curtailment': 0.0}))


if __name__ == "__main__":
    main()
